import openpyxl

location = ['常熟','吴江','太仓','张家港','昆山','姑苏区','相城区','吴中区','新区','园区']
# file_path = 'D:/安全审计/202404/202404/2、bes批量监控审计/苏州/苏州异常时间过户、开户、补卡、复机汇总20240401-0430.xlsx'
file_path = 'D:/安全审计/202404/202404/首三推查询日志202404.xlsx'
# 打开文件，获取工作簿
wb = openpyxl.load_workbook(file_path)
# print(wb.get_sheet_names())

# 获取工作簿中的工作表
sheet = wb.get_sheet_by_name('Sheet1')
# print(sheet.title)

# 修改单元格的值
# sheet['地市中文名'] = 'New Value'

max_col = sheet.max_column
sheet.cell(row=1,column=max_col+1,value='区县')
max_col += 1
for i in range(2,sheet.max_row+1):
    organization = sheet.cell(row=i,column=6).value
    flag = 0
    for loc in location:
        if organization.find(loc) != -1:
            # sheet.cell(row=i, column=max_col, value=loc)
            sheet.cell(row=i,column=max_col-1,value=loc)
            flag = 1
            break
    if flag == 0:
        sheet.cell(row=i, column=max_col-1, value='#N/A')

# 保存修改
wb.save(file_path)